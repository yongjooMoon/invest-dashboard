import streamlit as st
import requests
import urllib.parse
import json
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
from dateutil.relativedelta import relativedelta

# --- [페이지 설정] ---
st.set_page_config(page_title="통합 투자 분석기", page_icon="📈", layout="wide")

# --- [데이터 풀] ---
seoul_gu_pool = {
    '11110': '종로구', '11140': '중구', '11170': '용산구', '11200': '성동구', '11215': '광진구',
    '11230': '동대문구', '11260': '중랑구', '11290': '성북구', '11305': '강북구', '11320': '도봉구',
    '11350': '노원구', '11380': '은평구', '11410': '서대문구', '11440': '마포구', '11470': '양천구',
    '11500': '강서구', '11530': '구로구', '11545': '금천구', '11560': '영등포구', '11590': '동작구',
    '11620': '관악구', '11650': '서초구', '11680': '강남구', '11710': '송파구', '11740': '강동구'
}

seoul_dong_pool = {
    '11110': ['청운동', '신교동', '궁정동', '효자동', '창성동', '통의동', '적선동', '통인동', '누상동', '누하동', '옥인동', '체부동', '필운동', '내자동', '사직동',
              '도렴동', '당주동', '내수동', '세종로', '신문로1가', '신문로2가', '청진동', '서린동', '수송동', '중학동', '종로1가', '공평동', '관훈동', '견지동',
              '와룡동', '권농동', '운니동', '익선동', '경운동', '관철동', '인사동', '낙원동', '종로2가', '팔판동', '삼청동', '안국동', '소격동', '화동', '사간동',
              '송현동', '가회동', '재동', '평동', '평창동', '구기동', '부암동', '홍지동', '신영동', '무악동', '청운효자동', '창신동', '묘동', '돈의동', '종로3가',
              '관수동', '장사동', '종로4가', '예지동', '원남동', '종로5가', '종로6가', '이화동', '연건동', '충신동', '동숭동', '혜화동', '명륜1가', '명륜2가',
              '명륜4가', '명륜3가', '창경궁로', '창신1동', '창신2동', '창신3동', '숭인동', '교남동', '송월동', '홍파동', '교북동'],
    '11140': ['무교동', '다동', '태평로1가', '을지로1가', '태평로2가', '남대문로1가', '삼각동', '수하동', '장교동', '수표동', '소공동', '남창동', '북창동', '태평로',
              '남대문로', '봉래동', '회현동', '충무로', '명동', '남산동', '저동', '인현동', '예관동', '묵정동', '필동', '남학동', '주자동', '예장동', '장충동',
              '광희동', '쌍림동', '을지로', '주교동', '방산동', '오장동', '입정동', '산림동', '초동', '신당동', '흥인동', '무학동', '황학동', '서소문동', '정동',
              '순화동', '의주로', '중림동', '만리동'],
    '11170': ['후암동', '용산동', '갈월동', '남영동', '동자동', '서계동', '청파동', '원효로', '신창동', '산천동', '청암동', '효창동', '도원동', '용문동', '문배동',
              '신계동', '한강로', '이촌동', '이태원동', '한남동', '동빙고동', '서빙고동', '주성동', '용산동6가', '보광동'],
    '11200': ['상왕십리동', '하왕십리동', '홍익동', '도선동', '마장동', '사근동', '행당동', '응봉동', '금호동1가', '금호동2가', '금호동3가', '금호동4가', '옥수동',
              '성수동1가', '성수동2가', '송정동', '용답동'],
    '11215': ['중곡동', '능동', '구의동', '광장동', '자양동', '화양동', '군자동'],
    '11230': ['신설동', '용두동', '제기동', '전농동', '답십리동', '장안동', '청량리동', '회기동', '휘경동', '이문동'],
    '11260': ['면목동', '상봉동', '중화동', '묵동', '망우동', '신내동'],
    '11290': ['성북동', '돈암동', '동소문동', '삼선동', '안암동', '보문동', '정릉동', '길음동', '종암동', '하월곡동', '상월곡동', '장위동', '석관동'],
    '11305': ['미아동', '번동', '수유동', '우이동'],
    '11320': ['쌍문동', '미아동', '수유동', '창동', '도봉동', '방학동'],
    '11350': ['월계동', '공릉동', '하계동', '상계동', '중계동', '불암동'],
    '11380': ['수색동', '녹번동', '불광동', '갈현동', '구산동', '대조동', '응암동', '역촌동', '신사동', '증산동', '진관동'],
    '11410': ['충정로', '합동', '미근동', '냉천동', '천연동', '옥천동', '영천동', '현저동', '북아현동', '홍제동', '대신동', '신촌동', '봉원동', '창천동', '연희동',
              '홍은동', '북가좌동', '남가좌동'],
    '11440': ['아현동', '공덕동', '신공덕동', '도화동', '용강동', '토정동', '마포동', '대흥동', '염리동', '노고산동', '신수동', '현석동', '구수동', '창전동', '상수동',
              '하중동', '신정동', '당인동', '서교동', '동교동', '합정동', '망원동', '연남동', '성산동', '중동', '상암동'],
    '11470': ['목동', '신월동', '신정동'],
    '11500': ['염창동', '화곡동', '가양동', '마곡동', '내발산동', '외발산동', '방화동', '개화동', '과해동', '오곡동', '오쇠동'],
    '11530': ['신도림동', '구로동', '가리봉동', '고척동', '개봉동', '오류동', '궁동', '온수동', '천왕동', '항동'],
    '11545': ['가산동', '독산동', '시흥동'],
    '11560': ['영등포동', '여의도동', '당산동', '도림동', '문래동', '양평동', '양화동', '신길동', '대림동'],
    '11590': ['노량진동', '상도동', '본동', '흑석동', '동작동', '사당동', '대방동', '신대방동'],
    '11620': ['봉천동', '신림동', '남현동'],
    '11650': ['방배동', '양재동', '우면동', '원지동', '잠원동', '반포동', '서초동', '내곡동', '염곡동', '신원동'],
    '11680': ['역삼동', '개포동', '청담동', '삼성동', '대치동', '신사동', '논현동', '압구정동', '세곡동', '자곡동', '율현동', '일원동', '수서동', '도곡동'],
    '11710': ['잠실동', '신천동', '풍납동', '송파동', '석촌동', '삼전동', '가락동', '문정동', '장지동', '방이동', '오금동', '거여동', '마천동'],
    '11740': ['명일동', '고덕동', '상일동', '길동', '둔촌동', '암사동', '성내동', '천호동', '강일동']
}


# --- [함수부] ---
def clean_apt_name(name):
    if not name: return ""
    return name.replace(" ", "").replace("(주상복합)", "").replace("(도시형생활주택)", "").replace("주상복합", "").replace("아파트", "")


def extract_items(response_text):
    text = response_text.strip()
    if not text: return []
    if text.startswith('{'):
        try:
            data = json.loads(text)
            body = data.get('response', {}).get('body', {})
            if not isinstance(body, dict): return []

            if 'item' in body:
                items_node = body['item']
            else:
                items_parent = body.get('items', {})
                items_node = items_parent.get('item', []) if isinstance(items_parent, dict) else items_parent

            if isinstance(items_node, dict):
                return [items_node]
            elif isinstance(items_node, list):
                return items_node
            else:
                return []
        except:
            return []
    elif text.startswith('<'):
        try:
            root = ET.fromstring(text)
            return root.findall('.//item')
        except:
            return []
    return []


def get_field(item, *keys):
    if isinstance(item, dict):
        for k in keys:
            if k in item and item[k] is not None: return str(item[k]).strip()
        return ""
    else:
        for k in keys:
            val = item.findtext(k)
            if val is not None: return val.strip()
        return ""


def generate_excel_data(api_key, district_code, district_name, target_dong, start_date, end_date, apt_filters):
    log_text = ""
    yield "progress", "🔄 분석을 시작합니다..."

    api_key = urllib.parse.unquote(api_key.strip())
    url_apt_list_sigungu = 'https://apis.data.go.kr/1613000/AptListService3/getSigunguAptList3'
    url_apt_list_bjd = 'https://apis.data.go.kr/1613000/AptListService3/getLegaldongAptList3'
    url_apt_info = 'https://apis.data.go.kr/1613000/AptBasisInfoServiceV4/getAphusBassInfoV4'

    trade_api_urls = [
        'https://apis.data.go.kr/1613000/RTMSDataSvcAptTrade/getRTMSDataSvcAptTrade',
        'https://apis.data.go.kr/1613000/RTMSDataSvcAptPreSaleRightTrade/getRTMSDataSvcAptPreSaleRightTrade'
    ]

    kapt_name_to_code = {}
    apt_details_map = {}
    apt_types_master = {}
    deals_filtered_map = {}
    unique_bjd_codes = set()

    count_trade_api = 0
    count_kapt_api = 0

    master_scan_months = []
    current = datetime(start_date.year, start_date.month, 1)
    end_month_calc = datetime(end_date.year, end_date.month, 1)

    while current <= end_month_calc:
        master_scan_months.append(current.strftime("%Y%m"))
        current += relativedelta(months=1)

    dong_log = f"'{target_dong}' 전체" if target_dong != "전체" else "구 전체"
    log_text += f"\n🔄 [1단계] {master_scan_months[0]} ~ {master_scan_months[-1]} 기간 {dong_log} 실거래가 추출 중...\n"
    yield "log", log_text

    for deal_ymd in master_scan_months:
        for url_endpoint in trade_api_urls:
            trade_params = {'serviceKey': api_key, 'LAWD_CD': district_code, 'DEAL_YMD': deal_ymd, 'numOfRows': '10000'}
            try:
                res_trade = requests.get(url_endpoint, params=trade_params, timeout=10)
                count_trade_api += 1

                if "LIMITED NUMBER OF SERVICE REQUESTS EXCEEDS ERROR" in res_trade.text:
                    yield "error", "국토교통부 실거래가 API 일일 호출 한도(10,000회)를 초과했습니다."
                    return

                if res_trade.status_code == 200:
                    items = extract_items(res_trade.text)
                    for item in items:
                        try:
                            apt_nm = get_field(item, 'aptNm', '아파트명')
                            exclu_ar_str = get_field(item, 'excluUseAr', '전용면적')
                            dong = get_field(item, 'umdNm', '법정동')

                            if not apt_nm or not exclu_ar_str: continue
                            if target_dong != "전체" and target_dong not in dong: continue

                            if apt_filters:
                                clean_trade_apt = apt_nm.replace(" ", "")
                                matched = any(f.replace(" ", "") in clean_trade_apt for f in apt_filters)
                                if not matched: continue

                            s_code = get_field(item, 'sggCd', 'sigunguCd')
                            d_code = get_field(item, 'umdCd', 'eubmyundongCd')
                            if len(s_code) == 5 and len(d_code) == 5:
                                unique_bjd_codes.add(s_code + d_code)
                            elif len(d_code) == 5:
                                unique_bjd_codes.add(district_code + d_code)

                            exclu_ar = float(exclu_ar_str)
                            d_year = get_field(item, 'dealYear', '년')
                            d_month = get_field(item, 'dealMonth', '월').zfill(2)
                            d_day = get_field(item, 'dealDay', '일').zfill(2)

                            region_key = f"{district_name} ({dong})"
                            group_key = (region_key, apt_nm, exclu_ar)

                            apt_types_master[group_key] = get_field(item, 'buildYear', '건축년도')
                            deal_date_full = f"{d_year}.{d_month}.{d_day}"
                            amt_val = int(get_field(item, 'dealAmount', '거래금액').replace(',', '')) * 10

                            if group_key not in deals_filtered_map: deals_filtered_map[group_key] = []
                            deals_filtered_map[group_key].append({'deal_amount': amt_val, 'deal_date': deal_date_full})
                        except Exception:
                            continue
            except Exception:
                continue

    if not deals_filtered_map:
        yield "error", "설정하신 조건(기간/동/필터)에 해당하는 실거래 내역이 없습니다."
        return

    log_text += f"\n🏢 [2단계] K-APT 인덱싱 하이브리드 정밀 스캔 시작...\n"
    yield "log", log_text

    for bjd_code in unique_bjd_codes:
        list_params_bjd = {'serviceKey': api_key, 'bjdCode': bjd_code, 'numOfRows': '9999', 'pageNo': '1'}
        try:
            res_list = requests.get(url_apt_list_bjd, params=list_params_bjd, timeout=5)
            count_kapt_api += 1
            if "LIMITED NUMBER OF SERVICE REQUESTS EXCEEDS ERROR" in res_list.text:
                yield "error", "K-APT 단지목록 API 일일 한도를 초과했습니다."
                return
            for item in extract_items(res_list.text):
                kcode = get_field(item, 'kaptCode')
                if kcode: kapt_name_to_code[clean_apt_name(get_field(item, 'kaptName'))] = kcode
        except:
            pass

    list_params_sig = {'serviceKey': api_key, 'sigunguCode': district_code, 'numOfRows': '9999', 'pageNo': '1'}
    try:
        res_list = requests.get(url_apt_list_sigungu, params=list_params_sig, timeout=10)
        count_kapt_api += 1
        if "LIMITED NUMBER OF SERVICE REQUESTS EXCEEDS ERROR" in res_list.text:
            yield "error", "K-APT 시군구 API 일일 한도를 초과했습니다."
            return
        for item in extract_items(res_list.text):
            kcode = get_field(item, 'kaptCode')
            if kcode: kapt_name_to_code[clean_apt_name(get_field(item, 'kaptName'))] = kcode
    except:
        pass

    for apt_nm in set([key[1] for key in deals_filtered_map.keys()]):
        trade_cleaned_key = clean_apt_name(apt_nm)
        target_kcode = kapt_name_to_code.get(trade_cleaned_key)

        if not target_kcode:
            for k_clean, k_code in kapt_name_to_code.items():
                if trade_cleaned_key in k_clean or k_clean in trade_cleaned_key:
                    target_kcode = k_code;
                    break

        if target_kcode:
            try:
                info_params = {'serviceKey': api_key, 'kaptCode': target_kcode}
                res_info = requests.get(url_apt_info, params=info_params, timeout=5)
                count_kapt_api += 1

                if res_info.status_code == 200:
                    items = extract_items(res_info.text)
                    if items:
                        info_item = items[0]
                        raw_h_cnt = get_field(info_item, 'kaptdaCnt')
                        if not raw_h_cnt: raw_h_cnt = "-"

                        use_dt = get_field(info_item, 'kaptUsedate')
                        if not use_dt: use_dt = "-"

                        cls_type = get_field(info_item, 'codeAptNm')
                        classification = "주상복합" if ("주상복합" in cls_type or "스타클래스" in apt_nm) else (
                            "도시형생활주택" if "도시형" in apt_nm else "")

                        apt_details_map[trade_cleaned_key] = {'households': str(raw_h_cnt), 'move_in': str(use_dt),
                                                              'classification': classification}
                        log_text += f"✅ K-APT 매칭 완료: {apt_nm}\n"
                        yield "log", log_text
            except:
                pass
        else:
            log_text += f"❌ K-APT 매칭 실패: {apt_nm}\n"
            yield "log", log_text

        if trade_cleaned_key not in apt_details_map:
            classification = "주상복합" if "주상복합" in apt_nm or "스타클래스" in apt_nm else ""
            apt_details_map[trade_cleaned_key] = {'households': "-", 'move_in': "-", 'classification': classification}

    log_text += f"\n📊 [3단계] 엑셀 파일 생성 중...\n"
    yield "log", log_text

    sorted_keys = sorted(deals_filtered_map.keys(), key=lambda x: (x[0], x[1], x[2]))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "실거래 상세 분석"

    font_header = Font(name='맑은 고딕', size=10, bold=True)
    font_body = Font(name='맑은 고딕', size=10)
    font_summary = Font(name='맑은 고딕', size=10, bold=True, color='000000')
    fill_header = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    fill_summary = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_right = Alignment(horizontal='right', vertical='center')
    border_thin = Side(border_style="thin", color="D3D3D3")
    cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

    headers = ["구분", "단지명", "입주시기(원본)", "세대수(원본)", "공급면적(㎡)", "전용면적(㎡)", "공급(py)", "전용(py)", "전용률", "매매가", "공급평단가",
               "전용평단가", "계약일자"]
    ws.append(headers)

    for col_idx in range(1, 14):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = font_header;
        cell.fill = fill_header;
        cell.alignment = align_center;
        cell.border = cell_border

    current_row = 2

    for key in sorted_keys:
        region_text, apt_nm, exclu_ar = key
        trade_cleaned_key = clean_apt_name(apt_nm)
        display_apt_name = apt_nm
        info_p = apt_details_map.get(trade_cleaned_key, {})

        move_in_text = info_p.get('move_in', "-")
        household_text = info_p.get('households', "-")
        cls_type = info_p.get('classification', "")

        if cls_type == "주상복합" and "주상복합" not in display_apt_name:
            display_apt_name += "(주상복합)"
        elif cls_type == "도시형생활주택" and "도시형" not in display_apt_name:
            display_apt_name += "(도시형생활주택)"

        deals = deals_filtered_map[key]
        deals.sort(key=lambda x: x['deal_date'], reverse=True)

        for deal in deals:
            final_deal_amount = deal['deal_amount']
            final_deal_date = deal['deal_date']

            supply_m2 = round(exclu_ar / 0.76, 2)
            supply_py = round(supply_m2 * 0.3025, 2)
            exclu_py = round(exclu_ar * 0.3025, 2)
            ratio = (exclu_ar / supply_m2) if supply_m2 > 0 else 0
            supply_price = int(round(final_deal_amount / supply_py, 0)) if supply_py > 0 else 0
            exclu_price = int(round(final_deal_amount / exclu_py, 0)) if exclu_py > 0 else 0

            h_val = household_text
            if h_val.isdigit(): h_val = int(h_val)

            row_data = [region_text, display_apt_name, str(move_in_text), str(household_text), supply_m2, exclu_ar,
                        supply_py, exclu_py, ratio, final_deal_amount, supply_price, exclu_price, final_deal_date]
            ws.append(row_data)

            for col_idx in range(1, 14):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.font = font_body;
                cell.border = cell_border
                cell.alignment = align_center if col_idx in [1, 2, 3, 4, 13] else align_right
                if col_idx in [5, 6, 7, 8]:
                    cell.number_format = '#,##0.00'
                elif col_idx == 9:
                    cell.number_format = '0%'
                elif col_idx == 4 and isinstance(h_val, int):
                    cell.number_format = '#,##0'
                elif col_idx in [10, 11, 12]:
                    cell.number_format = '#,##0'

            current_row += 1

        amounts = [d['deal_amount'] for d in deals]
        avg_amt = sum(amounts) / len(amounts)
        supply_py = round((exclu_ar / 0.76) * 0.3025, 2)
        exclu_py = round(exclu_ar * 0.3025, 2)
        avg_supply_price = int(round(avg_amt / supply_py, 0)) if supply_py > 0 else 0
        avg_exclu_price = int(round(avg_amt / exclu_py, 0)) if exclu_py > 0 else 0

        summary_row = [
            region_text, display_apt_name, f"▶ {exclu_ar}㎡ 요약", None, None, None, None, None, None,
            f"최소: {min(amounts):,}\n최대: {max(amounts):,}\n평균: {int(avg_amt):,}",
            f"평균: {avg_supply_price:,}", f"평균: {avg_exclu_price:,}", None
        ]
        ws.append(summary_row)
        ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=9)

        for col_idx in range(1, 14):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.border = cell_border
            if col_idx >= 3:
                cell.fill = fill_summary;
                cell.font = font_summary
            else:
                cell.font = font_body
            if col_idx <= 3 or col_idx >= 10:
                cell.alignment = align_center
        current_row += 1

    merge_ranges = []
    start_r_A = 2
    for r in range(3, current_row + 1):
        if (ws.cell(row=r, column=1).value if r < current_row else None) != ws.cell(row=start_r_A, column=1).value:
            if r - 1 > start_r_A: merge_ranges.append((start_r_A, 1, r - 1, 1))
            start_r_A = r

    start_r_B = 2
    for r in range(3, current_row + 1):
        if (ws.cell(row=r, column=2).value if r < current_row else None) != ws.cell(row=start_r_B, column=2).value or \
                (ws.cell(row=r, column=1).value if r < current_row else None) != ws.cell(row=start_r_B, column=1).value:
            if r - 1 > start_r_B: merge_ranges.append((start_r_B, 2, r - 1, 2))
            start_r_B = r

    for m in merge_ranges:
        for r_idx in range(m[0] + 1, m[2] + 1): ws.cell(row=r_idx, column=m[1]).value = None
        ws.merge_cells(start_row=m[0], start_column=m[1], end_row=m[2], end_column=m[3])
        ws.cell(row=m[0], column=m[1]).alignment = align_center

    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 16

    ws.column_dimensions['A'].width = 18;
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 16;
    ws.column_dimensions['J'].width = 25
    ws.column_dimensions['K'].width = 18;
    ws.column_dimensions['L'].width = 18

    # 웹 다운로드를 위해 BytesIO에 엑셀 파일 저장
    output = BytesIO()
    wb.save(output)

    log_text += f"\n🎉 엑셀 대시보드 생성이 완료되었습니다!\n"
    log_text += f"🧾 이번 작업 API 사용량: 실거래가 {count_trade_api}회 / K-APT {count_kapt_api}회"
    yield "log", log_text

    yield "success", {
        "data": output.getvalue(),
        "filename": f"[{district_name}_{target_dong}]_실거래_분석.xlsx"
    }


# --- [Streamlit UI 프론트엔드] ---
# 사이드바 메뉴 구성
menu = st.sidebar.radio("메뉴 선택", ["🏢 부동산 실거래가 분석", "📈 주식 포트폴리오 (준비중)"])

if menu == "🏢 부동산 실거래가 분석":
    st.title("아파트 실거래가 정밀 분석기")
    st.markdown("공공데이터 API를 활용하여 부동산 실거래가를 엑셀로 자동 정리해 드립니다.")

    # 1. API 키 입력
    st.subheader("🔑 공공데이터 API Key")
    api_key = st.text_input("인증키를 입력하세요", type="password", help="data.go.kr에서 발급받은 Decoding 키를 입력하세요.")

    # 2. 지역 선택
    st.subheader("📍 대상 구 / 동 선택")
    col1, col2 = st.columns(2)
    with col1:
        gu_name = st.selectbox("자치구", list(seoul_gu_pool.values()), index=23)  # 기본값 송파구
    with col2:
        gu_code = [k for k, v in seoul_gu_pool.items() if v == gu_name][0]
        dong_list = ["전체 (구 단위)"] + sorted(seoul_dong_pool.get(gu_code, []))
        dong_name = st.selectbox("법정동", dong_list)

    # 3. 기간 선택
    st.subheader("📅 조회 기간")
    col3, col4 = st.columns(2)
    curr_year = datetime.now().year
    curr_month = datetime.now().month
    with col3:
        start_date = st.date_input("시작 연/월 (일자는 무시됩니다)", datetime(curr_year, 1, 1))
    with col4:
        end_date = st.date_input("종료 연/월 (일자는 무시됩니다)", datetime(curr_year, curr_month, 1))

    # 4. 필터 단어
    st.subheader("🏢 아파트명 필터")
    filter_text = st.text_input("검색할 단지명을 입력하세요 (쉼표로 구분, 비워두면 전체 스캔)", placeholder="예: 올림픽파크, 가락스타클래스")

    if st.button("✨ 실거래가 분석 시작", type="primary"):
        if not api_key:
            st.warning("API 인증키를 입력해주세요.")
            st.stop()
        if start_date > end_date:
            st.warning("시작월이 종료월보다 미래일 수 없습니다.")
            st.stop()

        apt_filters = [t.strip() for t in filter_text.split(',')] if filter_text.strip() else []
        if dong_name.startswith("전체") and not apt_filters:
            st.warning("구 전체를 스캔할 때는 최소 1개의 아파트 필터 단어가 필요합니다 (서버 과부하 방지).")
            st.stop()

        target_dong = "전체" if dong_name.startswith("전체") else dong_name

        # 로그창 생성
        log_area = st.empty()

        # 백그라운드 데이터 처리 실행
        result_excel = None
        for status, payload in generate_excel_data(api_key, gu_code, gu_name, target_dong, start_date, end_date,
                                                   apt_filters):
            if status == "progress":
                with st.spinner(payload):
                    pass
            elif status == "log":
                log_area.text_area("💻 실시간 데이터 분석 로그", payload, height=300)
            elif status == "error":
                st.error(f"🚫 오류 발생: {payload}")
                st.stop()
            elif status == "success":
                result_excel = payload

        if result_excel:
            st.success("엑셀 파일 생성이 완료되었습니다! 아래 버튼을 눌러 다운로드하세요.")
            st.download_button(
                label="📥 엑셀 파일 다운로드",
                data=result_excel["data"],
                file_name=result_excel["filename"],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

elif menu == "📈 주식 포트폴리오 (준비중)":
    st.title("주식 포트폴리오 분석")
    st.info("이 페이지는 현재 준비 중입니다. 다음 업데이트를 기대해 주세요!")